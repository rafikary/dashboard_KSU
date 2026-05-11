from flask import Blueprint, request, current_app, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import re
from datetime import datetime, timedelta
import json
import os

export_bp = Blueprint('export', __name__)

# ── Helper: Remove Emoticons ──────────────────────────────────────────
def remove_emoticons(text):
    """Hapus semua emoji/emoticon dari string"""
    if not isinstance(text, str):
        return text
    
    # Regex pattern untuk menghapus emoji
    emoji_pattern = re.compile("["
        u"\U0001F600-\U0001F64F"  # emoticons
        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
        u"\U0001F680-\U0001F6FF"  # transport & map symbols
        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
        u"\U00002702-\U000027B0"
        u"\U000024C2-\U0001F251"
        u"\U0001F900-\U0001F9FF"  # Supplemental Symbols and Pictographs
        u"\U0001FA00-\U0001FAFF"  # Chess Symbols
        "]+", flags=re.UNICODE)
    
    return emoji_pattern.sub(r'', text).strip()


# ── Helper: Format Excel ──────────────────────────────────────────────
def format_excel_header(ws, num_cols):
    """Format header dengan warna biru dan bold"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_adjust_columns(ws):
    """Auto-resize kolom based on content"""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def create_excel_from_dataframe(df, sheet_name="Data"):
    """Helper untuk convert DataFrame ke Excel di memory"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=remove_emoticons(str(header)))
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            clean_value = remove_emoticons(str(value)) if isinstance(value, str) else value
            ws.cell(row=row_idx, column=col_idx, value=clean_value)
    
    # Format
    format_excel_header(ws, len(headers))
    auto_adjust_columns(ws)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# ══════════════════════════════════════════════════════════════════════
# EXPORT ENDPOINTS
# ══════════════════════════════════════════════════════════════════════

# ── 1. TREND: DECLINE (Monthly) ───────────────────────────────────────
@export_bp.route('/trend/decline')
def export_trend_decline():
    """Export data customer dengan tren penjualan menurun (bulanan)"""
    df = current_app.config["DF_PENJUALAN"].copy()
    df["month"] = df["aritemdate"].dt.to_period("M").astype(str)
    
    monthly = (
        df.groupby(["custname", "month"])
        .agg(total=("aritemdtlamt", "sum"), orders=("aritemdate", "count"))
        .reset_index()
        .sort_values(["custname", "month"])
    )
    
    export_data = []
    for custname, group in monthly.groupby("custname"):
        if len(group) < 3:
            continue
        
        totals = group["total"].tolist()
        months_lst = group["month"].tolist()
        orders_lst = group["orders"].tolist()
        
        declines = sum(1 for i in range(len(totals) - 1) if totals[i] > totals[i + 1])
        pct_change = round(((totals[-1] - totals[-2]) / totals[-2]) * 100, 1) if totals[-2] else 0
        
        if declines >= 3:
            status = "Menurun"
        elif declines >= 2:
            status = "Waspada"
        else:
            status = "Stabil"
        
        export_data.append({
            "Customer": custname,
            "Status": status,
            "Perubahan (%)": pct_change,
            "Total Bulan": len(totals),
            "Bulan Menurun": declines,
            "Bulan Terakhir": months_lst[-1],
            "Revenue Terakhir": round(totals[-1], 0)
        })
    
    df_export = pd.DataFrame(export_data)
    df_export = df_export.sort_values("Perubahan (%)")
    
    excel_file = create_excel_from_dataframe(df_export, "Trend Decline")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'trend_decline_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# ── ACTIVITY PER CUSTOMER ─────────────────────────────────────────────
@export_bp.route('/trend/activity-per-customer')
def export_activity_per_customer():
    """Export customer activity data"""
    df = current_app.config["DF_PENJUALAN"].copy()
    limit = int(request.args.get("limit", 10))
    
    top_custs = (
        df.groupby("custname")["aritemdtlamt"]
        .sum()
        .nlargest(limit)
        .index.tolist()
    )
    
    df_top = df[df["custname"].isin(top_custs)].copy()
    df_top["month"] = df_top["aritemdate"].dt.to_period("M").astype(str)
    
    activity = (
        df_top.groupby(["custname", "month"])["aritemdate"]
        .count()
        .reset_index()
        .rename(columns={
            "custname": "Customer",
            "month": "Bulan",
            "aritemdate": "Jumlah Order"
        })
        .sort_values(["Customer", "Bulan"])
    )
    
    excel_file = create_excel_from_dataframe(activity, "Customer Activity")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'customer_activity_top{limit}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# ── 2. TREND: QoQ (Quarterly) ─────────────────────────────────────────
@export_bp.route('/trend/qoq')
def export_trend_qoq():
    """Export data customer dengan tren penjualan QoQ"""
    df = current_app.config["DF_PENJUALAN"].copy()
    df["quarter"] = df["aritemdate"].dt.to_period("Q").astype(str)
    
    qoq_matrix = (
        df.groupby(["custname", "quarter"])["aritemdtlamt"]
        .sum()
        .unstack(fill_value=0)
    )
    
    all_quarters = qoq_matrix.columns.tolist()
    
    export_data = []
    for cust, row in qoq_matrix.iterrows():
        totals = row.tolist()
        
        if sum(1 for t in totals if t > 0) < 2:
            continue
        
        last = totals[-1]
        prev = totals[-2]
        
        if prev > 0:
            pct = round(((last - prev) / prev) * 100, 1)
        else:
            pct = 100.0 if last > 0 else 0.0
        
        declines = sum(1 for i in range(len(totals)-1) if totals[i] > totals[i+1])
        
        if declines >= 3:
            status = "Menurun"
        elif declines >= 2:
            status = "Waspada"
        else:
            status = "Stabil"
        
        export_data.append({
            "Customer": cust,
            "Status": status,
            "Kuartal Terakhir": all_quarters[-1],
            "Revenue Terakhir": round(last, 0),
            "Kuartal Sebelumnya": all_quarters[-2],
            "Revenue Sebelumnya": round(prev, 0),
            "Perubahan QoQ (%)": pct,
            "Total Kuartal": len(totals),
            "Kuartal Menurun": declines
        })
    
    df_export = pd.DataFrame(export_data)
    df_export = df_export.sort_values("Perubahan QoQ (%)")
    
    excel_file = create_excel_from_dataframe(df_export, "Trend QoQ")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'trend_qoq_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 3. CUSTOMERS: INACTIVE ────────────────────────────────────────────
@export_bp.route('/customers/inactive')
def export_customers_inactive():
    """Export customer inactive"""
    df = current_app.config["DF_PENJUALAN"].copy()
    days = int(request.args.get("days", 90))
    
    now = df["aritemdate"].max()
    cutoff = now - timedelta(days=days)
    
    last_order = (
        df.groupby("custname")["aritemdate"]
        .max()
        .reset_index()
        .rename(columns={"aritemdate": "last_order_date"})
    )
    
    inactive_df = last_order[last_order["last_order_date"] < cutoff].copy()
    inactive_df["days_inactive"] = (now - inactive_df["last_order_date"]).dt.days
    inactive_df["last_order_date"] = inactive_df["last_order_date"].dt.strftime("%Y-%m-%d")
    inactive_df = inactive_df.sort_values("days_inactive", ascending=False)
    
    inactive_df = inactive_df.rename(columns={
        "custname": "Customer",
        "last_order_date": "Tanggal Order Terakhir",
        "days_inactive": "Hari Tidak Aktif"
    })
    
    excel_file = create_excel_from_dataframe(inactive_df, "Customer Inactive")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'customers_inactive_{days}days_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 4. CUSTOMERS: LOYALTY ─────────────────────────────────────────────
@export_bp.route('/customers/loyalty')
def export_customers_loyalty():
    """Export customer loyalty data"""
    df = current_app.config["DF_PENJUALAN"].copy()
    
    # Get period parameter (default: all_time)
    period = request.args.get("period", "all_time")
    
    # Filter by period if not all_time (month-based for accuracy)
    if period != "all_time":
        df['month_period'] = df['aritemdate'].dt.to_period('M')
        current_period = pd.Timestamp(datetime.now()).to_period("M")
        last_complete = current_period - 1
        data_max = df['month_period'].max()
        max_period = data_max if data_max < last_complete else last_complete
        
        if period == "3_months":
            start_period = max_period - 2  # Last 3 complete months
        elif period == "6_months":
            start_period = max_period - 5  # Last 6 complete months
        elif period == "12_months":
            start_period = max_period - 11  # Last 12 complete months
        else:
            start_period = max_period - 2  # default to 3 months
        
        df = df[(df['month_period'] >= start_period) & (df['month_period'] <= max_period)]
    
    loyalty_df = (
        df.groupby("custname")
        .agg(
            total_orders=("aritemdate", "count"),
            total_spend=("aritemdtlamt", "sum"),
            first_order=("aritemdate", "min"),
            last_order=("aritemdate", "max"),
        )
        .reset_index()
        .sort_values("total_spend", ascending=False)
    )
    
    loyalty_df["first_order"] = loyalty_df["first_order"].dt.strftime("%Y-%m-%d")
    loyalty_df["last_order"] = loyalty_df["last_order"].dt.strftime("%Y-%m-%d")
    loyalty_df["total_spend"] = loyalty_df["total_spend"].round(0)
    
    loyalty_df = loyalty_df.rename(columns={
        "custname": "Customer",
        "total_orders": "Total Transaksi",
        "total_spend": "Total Belanja",
        "first_order": "Order Pertama",
        "last_order": "Order Terakhir"
    })
    
    # Add period info to filename
    period_labels = {
        'all_time': 'all_time',
        '12_months': '12months',
        '6_months': '6months',
        '3_months': '3months'
    }
    period_label = period_labels.get(period, 'all_time')
    
    excel_file = create_excel_from_dataframe(loyalty_df, "Customer Loyalty")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'customers_loyalty_{period_label}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 5. CUSTOMERS: TOP SPENDING ────────────────────────────────────────
@export_bp.route('/customers/spending')
def export_customers_spending():
    """Export top spending customers"""
    df = current_app.config["DF_PENJUALAN"].copy()
    month = request.args.get("month")
    
    if month:
        df = df[df["aritemdate"].dt.to_period("M").astype(str) == month]
    
    result = (
        df.groupby("custname")["aritemdtlamt"]
        .sum()
        .reset_index()
        .rename(columns={"custname": "Customer", "aritemdtlamt": "Total Belanja"})
        .sort_values("Total Belanja", ascending=False)
    )
    result["Total Belanja"] = result["Total Belanja"].round(0)
    
    excel_file = create_excel_from_dataframe(result, "Top Spending")
    
    filename = f'top_spending_{month}_{datetime.now().strftime("%Y%m%d")}.xlsx' if month else f'top_spending_alltime_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── 6. ITEMS: TOP ITEMS ───────────────────────────────────────────────
@export_bp.route('/items/top-overall')
def export_top_items():
    """Export top items by revenue with month range filter"""
    df = current_app.config["DF_PENJUALAN"].copy()
    from_month = request.args.get("from")
    to_month = request.args.get("to")
    limit = int(request.args.get("limit", 50))
    
    # Filter by month range if provided
    if from_month:
        df = df[df["aritemdate"].dt.to_period("M").astype(str) >= from_month]
    if to_month:
        df = df[df["aritemdate"].dt.to_period("M").astype(str) <= to_month]
    
    top_items = (
        df.groupby("itemshortdesc")
        .agg(
            total_revenue=("aritemdtlamt", "sum"),
            total_qty=("aritemqty", "sum"),
            total_orders=("aritemdate", "count")
        )
        .reset_index()
        .sort_values("total_revenue", ascending=False)
        .head(limit)
    )
    
    top_items = top_items.rename(columns={
        "itemshortdesc": "Item",
        "total_revenue": "Total Revenue",
        "total_qty": "Total Quantity",
        "total_orders": "Total Transaksi"
    })
    
    excel_file = create_excel_from_dataframe(top_items, "Top Items")
    
    period_str = f"{from_month}_to_{to_month}" if from_month and to_month else "alltime"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'top_items_{period_str}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 7. ITEMS: TOP BY CUSTOMER ─────────────────────────────────────────
@export_bp.route('/items/top-by-customer/<customer_name>')
def export_top_items_by_customer(customer_name):
    """Export top items untuk customer tertentu (all-time)"""
    df = current_app.config["DF_PENJUALAN"].copy()
    limit = int(request.args.get("limit", 50))
    
    # Filter customer (case insensitive) - no time filter, all-time
    df_cust = df[df["custname"].str.upper() == customer_name.upper()]
    
    if df_cust.empty:
        # Return empty file
        df_export = pd.DataFrame(columns=["Item", "Total Revenue", "Total Quantity", "Total Transaksi"])
    else:
        top_items = (
            df_cust.groupby("itemshortdesc")
            .agg(
                total_revenue=("aritemdtlamt", "sum"),
                total_qty=("aritemqty", "sum"),
                total_orders=("aritemdate", "count")
            )
            .reset_index()
            .sort_values("total_revenue", ascending=False)
            .head(limit)
        )
        
        df_export = top_items.rename(columns={
            "itemshortdesc": "Item",
            "total_revenue": "Total Revenue",
            "total_qty": "Total Quantity",
            "total_orders": "Total Transaksi"
        })
    
    excel_file = create_excel_from_dataframe(df_export, "Top Items")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'top_items_{customer_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 8. GEOGRAPHY: AREA RANKING ────────────────────────────────────────
@export_bp.route('/geography/ranking')
def export_geography_ranking():
    """Export ranking customer per area with month range filter"""
    df = current_app.config["DF_PENJUALAN"].copy()
    from_month = request.args.get("from")
    to_month = request.args.get("to")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    
    # Prioritaskan date_from/date_to (full date), fallback ke from/to (month string)
    if date_from:
        df = df[df["aritemdate"] >= pd.to_datetime(date_from)]
    elif from_month:
        df = df[df["aritemdate"].dt.to_period("M").astype(str) >= from_month]
    if date_to:
        df = df[df["aritemdate"] < pd.to_datetime(date_to) + pd.Timedelta(days=1)]
    elif to_month:
        df = df[df["aritemdate"].dt.to_period("M").astype(str) <= to_month]
    
    ranking = df.groupby(['area', 'custname'])['aritemdtlamt'].sum().reset_index()
    ranking.columns = ['Area', 'Customer', 'Total Omzet']
    ranking = ranking.sort_values(['Area', 'Total Omzet'], ascending=[True, False])
    
    excel_file = create_excel_from_dataframe(ranking, "Ranking by Area")
    
    period_str = f"{date_from}_to_{date_to}" if date_from and date_to else (f"{from_month}_to_{to_month}" if from_month and to_month else "alltime")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'geography_ranking_{period_str}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 9. CATEGORY: BREAKDOWN ────────────────────────────────────────────
@export_bp.route('/category/breakdown')
def export_category_breakdown():
    """Export category breakdown"""
    df = current_app.config["DF_PENJUALAN"].copy()
    level = request.args.get('level', 'cat1')
    
    cat_column = 'cat1shortdesc' if level == 'cat1' else 'cat2shortdesc'
    breakdown = df.groupby(cat_column)['aritemdtlamt'].sum().reset_index()
    breakdown.columns = ['Kategori', 'Total Omzet']
    
    total_omzet = breakdown['Total Omzet'].sum()
    breakdown['Persentase (%)'] = (breakdown['Total Omzet'] / total_omzet * 100).round(2)
    breakdown = breakdown.sort_values('Total Omzet', ascending=False)
    
    excel_file = create_excel_from_dataframe(breakdown, "Category Breakdown")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'category_breakdown_{level}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 9B. CATEGORY: TRENDS ───────────────────────────────────────────────
@export_bp.route('/category/trends')
def export_category_trends():
    """Export category trends over time"""
    df = current_app.config["DF_PENJUALAN"].copy()
    level = request.args.get('level', 'cat1')
    period = request.args.get('period', 'quarterly')
    months = int(request.args.get('months', 12))
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Filter date — prioritaskan date_from/date_to, lalu months
    if date_from or date_to:
        if date_from:
            df = df[df['aritemdate'] >= pd.to_datetime(date_from)].copy()
        if date_to:
            df = df[df['aritemdate'] <= pd.to_datetime(date_to)].copy()
    else:
        latest_date = df['aritemdate'].max()
        cutoff_date = latest_date - timedelta(days=30 * months)
        df = df[df['aritemdate'] >= cutoff_date].copy()
    
    # Create period column
    if period == 'monthly':
        df['period'] = df['aritemdate'].dt.to_period('M').astype(str)
    else:
        df['period'] = df['aritemdate'].dt.to_period('Q').astype(str)
    
    # Aggregate
    cat_column = 'cat1shortdesc' if level == 'cat1' else 'cat2shortdesc'
    trends = df.groupby([cat_column, 'period'])['aritemdtlamt'].sum().reset_index()
    
    # Pivot to wide format
    trends_pivot = trends.pivot(index=cat_column, columns='period', values='aritemdtlamt').fillna(0)
    trends_pivot.columns.name = None
    trends_pivot.index.name = 'Kategori'
    trends_pivot = trends_pivot.reset_index()
    
    # Sort by total descending
    trends_pivot['Total'] = trends_pivot.iloc[:, 1:].sum(axis=1)
    trends_pivot = trends_pivot.sort_values('Total', ascending=False).drop(columns='Total')
    
    excel_file = create_excel_from_dataframe(trends_pivot, "Category Trends")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'category_trends_{period}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 10. MARKETING: PERFORMANCE ────────────────────────────────────────
@export_bp.route('/marketing/performance')
def export_marketing_performance():
    """Export marketing performance"""
    from routes.marketing import get_df_with_marketing, filter_df_by_period

    months    = request.args.get('months', 6)
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    df = get_df_with_marketing()

    df_r, _ = filter_df_by_period(df, months, date_from, date_to)

    result = []
    dup_mask_r = df_r.duplicated(keep=False)
    neg_mask_r = df_r['aritemqty'] < 0

    for person_name in df_r['marketing_person'].unique():
        df_p     = df_r[df_r['marketing_person'] == person_name]
        df_p_dup = df_r[dup_mask_r & (df_r['marketing_person'] == person_name)]
        df_p_neg = df_r[neg_mask_r & (df_r['marketing_person'] == person_name)]

        total_omzet    = float(df_p['aritemdtlamt'].sum())
        total_cust     = int(df_p['custname'].nunique())
        total_tx       = int(len(df_p))
        dup_omzet      = float(df_p_dup['aritemdtlamt'].sum())
        void_omzet     = float(df_p_neg['aritemdtlamt'].sum())
        dirty_tx       = int(len(df_p_dup)) + int(len(df_p_neg))
        integrity_rate = round((total_tx - dirty_tx) / total_tx * 100, 1) if total_tx > 0 else 100.0
        clean_omzet    = total_omzet - dup_omzet - abs(void_omzet)

        result.append({
            'Marketing':       person_name,
            'Total Omzet':     total_omzet,
            'Clean Omzet':     clean_omzet,
            'Duplikat Omzet':  dup_omzet,
            'Void Omzet':      abs(void_omzet),
            'Integrity Rate (%)': integrity_rate,
            'Total Customer':  total_cust,
            'Total Transaksi': total_tx
        })

    df_export = pd.DataFrame(result).sort_values('Clean Omzet', ascending=False)
    excel_file = create_excel_from_dataframe(df_export, "Marketing Performance")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'marketing_performance_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 11. DATA QUALITY: SUMMARY ─────────────────────────────────────────
@export_bp.route('/dataquality/summary')
def export_dataquality():
    """Export data quality summary"""
    df = current_app.config['DF_PENJUALAN'].copy()
    df['custname'] = df['custname'].str.strip()
    df['itemshortdesc'] = df['itemshortdesc'].str.strip()
    
    total_rows = len(df)
    
    # Duplikat
    dup_mask = df.duplicated(keep=False)
    dup_rows = int(dup_mask.sum())
    dup_first_mask = df.duplicated(keep='first')
    dup_first = int(dup_first_mask.sum())
    
    # Qty negatif
    neg_mask = df['aritemqty'] < 0
    neg_rows = int(neg_mask.sum())
    neg_amount = float(df[neg_mask]['aritemdtlamt'].sum())
    
    # Amount nol
    zero_mask = df['aritemdtlamt'] == 0
    zero_rows = int(zero_mask.sum())
    
    # Clean rows
    clean_df = df[~dup_first_mask & ~neg_mask & ~zero_mask]
    clean_rows = len(clean_df)
    clean_omzet = float(clean_df['aritemdtlamt'].sum())
    dirty_omzet = float(df['aritemdtlamt'].sum())
    
    summary_data = [{
        'Metric': 'Total Rows',
        'Value': total_rows
    }, {
        'Metric': 'Clean Rows',
        'Value': clean_rows
    }, {
        'Metric': 'Dirty Rows',
        'Value': total_rows - clean_rows
    }, {
        'Metric': 'Dirty Percentage (%)',
        'Value': round((total_rows - clean_rows) / total_rows * 100, 2)
    }, {
        'Metric': 'Duplicate Rows',
        'Value': dup_rows
    }, {
        'Metric': 'Removable Duplicates',
        'Value': dup_first
    }, {
        'Metric': 'Negative Qty Rows',
        'Value': neg_rows
    }, {
        'Metric': 'Zero Amount Rows',
        'Value': zero_rows
    }, {
        'Metric': 'Raw Omzet',
        'Value': dirty_omzet
    }, {
        'Metric': 'Clean Omzet',
        'Value': clean_omzet
    }]
    
    df_export = pd.DataFrame(summary_data)
    
    excel_file = create_excel_from_dataframe(df_export, "Data Quality")
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'data_quality_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ── 12. MONTHLY REVENUE ───────────────────────────────────────────────
@export_bp.route('/trend/monthly-revenue')
def export_monthly_revenue():
    """Export monthly revenue data"""
    df = current_app.config["DF_PENJUALAN"].copy()
    area = request.args.get("area")
    
    if area:
        df = df[df["area"].str.upper() == area.upper()]
    
    monthly = (
        df.groupby(df["aritemdate"].dt.to_period("M"))
        .agg(total_revenue=("aritemdtlamt", "sum"), total_orders=("aritemdate", "count"))
        .reset_index()
    )
    monthly["month"] = monthly["aritemdate"].astype(str)
    monthly["total_revenue"] = monthly["total_revenue"].round(0)
    
    df_export = monthly[["month", "total_revenue", "total_orders"]].rename(columns={
        "month": "Bulan",
        "total_revenue": "Total Revenue",
        "total_orders": "Total Transaksi"
    })
    
    excel_file = create_excel_from_dataframe(df_export, "Monthly Revenue")
    
    filename = f'monthly_revenue_{area}_{datetime.now().strftime("%Y%m%d")}.xlsx' if area else f'monthly_revenue_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── 13. AI REPORT ──────────────────────────────────────────────────────
def smart_filter_dataframe(df, column, search_term):
    """
    Smart filtering with multiple fallback strategies:
    1. Exact match (case-insensitive)
    2. Partial match with escaped special characters
    3. Fuzzy multi-word match
    """
    # Clean the column
    df[column] = df[column].astype(str).str.strip()
    
    # Strategy 1: Exact match
    exact_mask = df[column].str.upper() == search_term.upper()
    if exact_mask.any():
        return df[exact_mask].copy()
    
    # Strategy 2: Partial match (escape special regex chars)
    try:
        safe_term = re.escape(search_term)
        contains_mask = df[column].str.contains(safe_term, case=False, na=False, regex=True)
        if contains_mask.any():
            return df[contains_mask].copy()
    except Exception as e:
        print(f"[Filter] Regex error: {e}")
    
    # Strategy 3: Fuzzy multi-word match
    words = search_term.split()
    if len(words) > 1:
        fuzzy_mask = pd.Series(True, index=df.index)
        for word in words:
            if len(word) > 2:  # Only consider words longer than 2 chars
                safe_word = re.escape(word)
                fuzzy_mask = fuzzy_mask & df[column].str.contains(safe_word, case=False, na=False, regex=True)
        
        if fuzzy_mask.any():
            return df[fuzzy_mask].copy()
    
    # Nothing found
    return pd.DataFrame()


@export_bp.route('/aireport')
def export_ai_report():
    """Export AI Report data to Excel"""
    try:
        df = current_app.config["DF_PENJUALAN"].copy()
        
        report_type = request.args.get("type", "general")
        entity_id = request.args.get("id", "")
        
        print(f"[AI Report Export] type={report_type}, id={entity_id}")
        
        # Ensure datetime format
        if not pd.api.types.is_datetime64_any_dtype(df['aritemdate']):
            df['aritemdate'] = pd.to_datetime(df['aritemdate'])
        
        # Filter data based on type and entity using smart filtering
        if report_type == "customer" and entity_id:
            filtered_df = smart_filter_dataframe(df, 'custname', entity_id)
            entity_label = "Customer"
        elif report_type == "product" and entity_id:
            filtered_df = smart_filter_dataframe(df, 'itemshortdesc', entity_id)
            entity_label = "Product"
        elif report_type == "sales" and entity_id:
            if 'salesperson' in df.columns:
                filtered_df = smart_filter_dataframe(df, 'salesperson', entity_id)
            elif 'salesoid' in df.columns:
                filtered_df = smart_filter_dataframe(df, 'salesoid', entity_id)
            else:
                filtered_df = df.copy()
            entity_label = "Sales Person"
        else:
            # General report (all data)
            filtered_df = df.copy()
            entity_label = "All Entities"
        
        if filtered_df.empty:
            return jsonify({
                "error": True,
                "message": f"No data found for {entity_label}: {entity_id}"
            }), 404
        
        # Prepare export dataframe
        export_cols = ['aritemdate', 'custname', 'itemshortdesc', 'aritemqty', 'aritemdtlamt']
        
        # Add additional columns if they exist
        if 'cat1shortdesc' in filtered_df.columns:
            export_cols.insert(3, 'cat1shortdesc')
        if 'area' in filtered_df.columns:
            export_cols.insert(2, 'area')
        
        # Filter only existing columns
        export_cols = [col for col in export_cols if col in filtered_df.columns]
        
        export_df = filtered_df[export_cols].copy()
        
        # Format date
        export_df['aritemdate'] = pd.to_datetime(export_df['aritemdate']).dt.strftime('%Y-%m-%d')
        
        # Round amounts
        if 'aritemdtlamt' in export_df.columns:
            export_df['aritemdtlamt'] = export_df['aritemdtlamt'].round(0)
        
        # Rename columns to Indonesian
        column_mapping = {
            'aritemdate': 'Tanggal',
            'custname': 'Nama Customer',
            'area': 'Area',
            'itemshortdesc': 'Nama Barang',
            'cat1shortdesc': 'Kategori',
            'aritemqty': 'Jumlah',
            'aritemdtlamt': 'Total Nilai'
        }
        
        export_df = export_df.rename(columns=column_mapping)
        
        # Sort by date descending
        export_df = export_df.sort_values('Tanggal', ascending=False)
        
        # Create Excel file
        excel_file = create_excel_from_dataframe(export_df, "AI Report Data")
        
        # Generate filename
        safe_entity = re.sub(r'[^\w\s-]', '', entity_id).strip().replace(' ', '_')[:50]
        filename = f'ai_report_{report_type}_{safe_entity}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"[AI Report Export ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": True,
            "message": f"Export failed: {str(e)}"
        }), 500
